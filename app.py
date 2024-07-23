from flask import Flask, Blueprint, request, render_template, redirect, url_for, flash, session
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook
import pandas as pd
import os

app = Flask(__name__)
db = SQLAlchemy()

# SQLite 데이터베이스 경로 설정
basedir = os.path.abspath(os.path.dirname(__file__))
db_path = os.path.join(basedir, 'instance', 'users.db')
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'your_secret_key'

db.init_app(app)

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.String(150), unique=True, nullable=False)
    user_pw = db.Column(db.String(150), nullable=False)
    student_grade = db.Column(db.String(50), nullable=False)
    student_class = db.Column(db.String(50), nullable=True)
    student_number = db.Column(db.String(50), nullable=True)
    barcode = db.Column(db.String(50), nullable=True)
    role = db.Column(db.String(50), nullable=False)
    name = db.Column(db.String(150), nullable=False)

class OutingRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, nullable=False)
    student_name = db.Column(db.String(150), nullable=False)
    student_class = db.Column(db.String(50), nullable=False)
    barcode = db.Column(db.String(50), nullable=False)
    out_time = db.Column(db.String(50), nullable=False)
    in_time = db.Column(db.String(50), nullable=False)
    approved = db.Column(db.Boolean, default=False)

# 데이터베이스 파일이 없을 때만 생성
if not os.path.exists(db_path):
    with app.app_context():
        db.create_all()

excel_file_path = 'students.xlsx'

def load_students():
    """엑셀 파일을 읽고 데이터프레임을 반환합니다."""
    return pd.read_excel(excel_file_path, dtype={'barcode': str})

df = load_students()

register_bp = Blueprint('register', __name__)
login_bp = Blueprint('login', __name__)
outing_bp = Blueprint('outing', __name__)

@register_bp.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        user_id = request.form.get('username')
        user_pw = request.form.get('password')
        role = request.form.get('role')
        name = request.form.get('student_name') if role == 'student' else request.form.get('teacher_name')
        student_grade = request.form.get('grade') if role == 'student' else request.form.get('teacher_grade')
        student_class = request.form.get('class') if role == 'student' and student_grade != '기타' else request.form.get('teacher_class')
        student_number = request.form.get('student_id') if role == 'student' else None
        barcode = request.form.get('barcode') if role == 'student' else None

        if role == 'student' and not barcode:
            flash('학생은 바코드 정보를 입력해야 합니다.', 'error')
            return redirect(url_for('register.register'))

        existing_user = User.query.filter_by(user_id=user_id).first()
        if existing_user:
            flash('이미 존재하는 사용자 이름입니다.', 'error')
            return redirect(url_for('register.register'))

        hashed_password = generate_password_hash(user_pw, method='pbkdf2:sha256')

        new_user = User(
            user_id=user_id,
            user_pw=hashed_password,
            student_class=student_class,
            role=role,
            barcode=barcode,
            name=name,
            student_grade=student_grade,
            student_number=student_number
        )

        db.session.add(new_user)
        db.session.commit()

        flash('회원가입이 완료되었습니다. 이제 로그인하세요.', 'success')
        return redirect(url_for('login.login'))

    return render_template('register.html')

@login_bp.route('/login/student', methods=['GET', 'POST'])
def login_student():
    if request.method == 'POST':
        user_id = request.form.get('username')
        user_pw = request.form.get('password')
        user = User.query.filter_by(user_id=user_id, role='student').first()

        if user and check_password_hash(user.user_pw, user_pw):
            session['user_id'] = user.user_id
            session['user_class'] = user.student_class
            session['role'] = user.role

            # 세션 데이터 출력
            print(f"Session Data: user_id={session['user_id']}, user_class={session['user_class']}, role={session['role']}")

            return redirect(url_for('outing.student_home'))

        flash('아이디 또는 비밀번호가 올바르지 않습니다.', 'error')
        return redirect(url_for('login.login_student'))

    return render_template('login.html')

@login_bp.route('/login/teacher', methods=['GET', 'POST'])
def login_teacher():
    if request.method == 'POST':
        user_id = request.form.get('username')
        user_pw = request.form.get('password')
        user = User.query.filter_by(user_id=user_id, role='teacher').first()

        if user and check_password_hash(user.user_pw, user_pw):
            session['user_id'] = user.user_id
            session['user_class'] = user.student_class
            session['role'] = user.role

            # 세션 데이터 출력
            print(f"Session Data: user_id={session['user_id']}, user_class={session['user_class']}, role={session['role']}")

            return redirect(url_for('outing.class_page', class_name=user.student_class))

        flash('아이디 또는 비밀번호가 올바르지 않습니다.', 'error')
        return redirect(url_for('login.login_teacher'))

    return render_template('login.html')

@login_bp.route('/logout')
def logout():
    session.pop('user_id', None)
    session.pop('user_class', None)
    session.pop('role', None)
    return redirect(url_for('login.login_student'))

@login_bp.route('/', methods=['GET'])
def home_page():
    if 'user_id' in session:
        # 세션 데이터 출력
        print(f"Home Page Session Data: user_id={session['user_id']}, user_class={session['user_class']}, role={session['role']}")

        if session['role'] == 'student':
            return redirect(url_for('outing.student_home'))
        else:
            return redirect(url_for('outing.class_page', class_name=session['user_class']))
    return redirect(url_for('login.login_student'))

@outing_bp.route('/student_home', methods=['GET'])
def student_home():
    if 'user_id' not in session or session['role'] != 'student':
        return redirect(url_for('login.login_student'))

    student = User.query.filter_by(user_id=session['user_id']).first()
    return render_template('student_home.html', student=student)

@outing_bp.route('/request_outing', methods=['POST'])
def request_outing():
    if 'user_id' not in session or session['role'] != 'student':
        return redirect(url_for('login.login_student'))

    student = User.query.filter_by(user_id=session['user_id']).first()
    out_time = request.form.get('out_time').replace('T', ' ')
    in_time = request.form.get('in_time').replace('T', ' ')

    outing_request = OutingRequest(
        student_id=student.id,
        student_name=student.user_id,
        student_class=student.student_class,
        barcode=student.barcode,
        out_time=out_time,
        in_time=in_time
    )

    db.session.add(outing_request)
    db.session.commit()

    flash('외출 신청이 완료되었습니다.', 'success')
    return redirect(url_for('outing.student_home'))

@outing_bp.route('/class/<class_name>', methods=['GET', 'POST'])
def class_page(class_name):
    global df  # global 선언이 df 변수 사용 전에 위치하도록 이동

    if 'user_id' not in session or session['role'] != 'teacher':
        return redirect(url_for('login.login_teacher'))

    filtered_df = df[df['class'] == class_name]
    if request.method == 'POST':
        barcode = request.form.get('barcode')
        out_datetime = request.form.get('out_datetime').replace('T', ' ')
        in_datetime = request.form.get('in_datetime').replace('T', ' ')

        # 엑셀 파일 수정
        wb = load_workbook(excel_file_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == barcode:
                row[2].value = out_datetime
                row[3].value = in_datetime
                break

        # 엑셀 파일 저장
        wb.save(excel_file_path)

        # 데이터프레임 갱신
        df = load_students()

        return redirect(url_for('outing.class_page', class_name=class_name))

    return render_template('class.html', students=filtered_df.to_dict(orient='records'), class_name=class_name)

@outing_bp.route('/manage_requests', methods=['GET'])
def manage_requests():
    if 'user_id' not in session or session['role'] != 'teacher':
        return redirect(url_for('login.login_teacher'))

    user_class = session['user_class']
    requests = OutingRequest.query.filter_by(student_class=user_class, approved=False).all()
    return render_template('manage_requests.html', requests=requests)

@outing_bp.route('/approve_request/<int:request_id>', methods=['POST'])
def approve_request(request_id):
    global df  # global 선언이 df 변수 사용 전에 위치하도록 이동

    if 'user_id' not in session or session['role'] != 'teacher':
        return redirect(url_for('login.login_teacher'))

    outing_request = OutingRequest.query.get(request_id)
    if outing_request:
        outing_request.approved = True
        db.session.commit()

        # 엑셀 파일 수정
        wb = load_workbook(excel_file_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == outing_request.barcode:
                row[2].value = outing_request.out_time
                row[3].value = outing_request.in_time
                break

        # 엑셀 파일 저장
        wb.save(excel_file_path)

        # 데이터프레임 갱신
        df = load_students()

    return redirect(url_for('outing.manage_requests'))

app.register_blueprint(register_bp)
app.register_blueprint(login_bp, url_prefix='/')
app.register_blueprint(outing_bp)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)
