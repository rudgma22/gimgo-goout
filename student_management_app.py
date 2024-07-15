from flask import Flask, request, render_template, redirect, url_for, session, flash
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook
import pandas as pd
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime

# 엑셀 파일 경로
excel_file_path = 'students.xlsx'


def load_students():
    """엑셀 파일을 읽고 데이터프레임을 반환합니다."""
    return pd.read_excel(excel_file_path, dtype={'barcode': str})


# 엑셀 파일 읽기
df = load_students()

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///users.db'
app.config['SECRET_KEY'] = 'your_secret_key'
db = SQLAlchemy(app)


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), unique=True, nullable=False)
    password = db.Column(db.String(150), nullable=False)
    user_class = db.Column(db.String(50), nullable=False)
    barcode = db.Column(db.String(50), nullable=True)  # 바코드는 학생에게만 해당
    role = db.Column(db.String(50), nullable=False)  # 'student' 또는 'teacher'


class OutingRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, nullable=False)
    student_name = db.Column(db.String(150), nullable=False)
    student_class = db.Column(db.String(50), nullable=False)
    barcode = db.Column(db.String(50), nullable=False)
    out_time = db.Column(db.String(50), nullable=False)
    in_time = db.Column(db.String(50), nullable=False)
    approved = db.Column(db.Boolean, default=False)


with app.app_context():
    db.create_all()


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user_class = request.form.get('class')
        role = request.form.get('role')
        barcode = request.form.get('barcode')

        if role == 'student' and not barcode:
            flash('학생은 바코드 정보를 입력해야 합니다.', 'error')
            return redirect(url_for('register'))

        hashed_password = generate_password_hash(password, method='pbkdf2:sha256')

        new_user = User(username=username, password=hashed_password, user_class=user_class, role=role, barcode=barcode)
        db.session.add(new_user)
        db.session.commit()

        flash('회원가입이 완료되었습니다. 이제 로그인하세요.', 'success')
        return redirect(url_for('login'))

    return render_template('register.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role')
        user = User.query.filter_by(username=username, role=role).first()

        if user and check_password_hash(user.password, password):
            session['username'] = user.username
            session['user_class'] = user.user_class
            session['role'] = user.role

            if role == 'student':
                return redirect(url_for('student_home'))
            else:
                return redirect(url_for('class_page', class_name=user.user_class))

        flash('아이디 또는 비밀번호가 올바르지 않습니다.', 'error')
        return redirect(url_for('login'))

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.pop('username', None)
    session.pop('user_class', None)
    session.pop('role', None)
    return redirect(url_for('login'))


@app.route('/', methods=['GET'])
def home_page():
    if 'username' in session:
        if session['role'] == 'student':
            return redirect(url_for('student_home'))
        else:
            return redirect(url_for('class_page', class_name=session['user_class']))
    return redirect(url_for('login'))


@app.route('/student_home', methods=['GET'])
def student_home():
    if 'username' not in session or session['role'] != 'student':
        return redirect(url_for('login'))

    student = User.query.filter_by(username=session['username']).first()
    return render_template('student_home.html', student=student)


@app.route('/request_outing', methods=['POST'])
def request_outing():
    if 'username' not in session or session['role'] != 'student':
        return redirect(url_for('login'))

    student = User.query.filter_by(username=session['username']).first()
    out_time = request.form.get('out_time').replace('T', ' ')
    in_time = request.form.get('in_time').replace('T', ' ')

    outing_request = OutingRequest(
        student_id=student.id,
        student_name=student.username,
        student_class=student.user_class,
        barcode=student.barcode,
        out_time=out_time,
        in_time=in_time
    )

    db.session.add(outing_request)
    db.session.commit()

    flash('외출 신청이 완료되었습니다.', 'success')
    return redirect(url_for('student_home'))


@app.route('/class/<class_name>', methods=['GET', 'POST'])
def class_page(class_name):
    if 'username' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    filtered_df = df[df['class'] == class_name]
    if request.method == 'POST':
        barcode = request.form.get('barcode')
        out_datetime = request.form.get('out_datetime').replace('T', ' ')
        in_datetime = request.form.get('in_datetime').replace('T', ' ')

        # 엑셀 파일 수정
        wb = load_workbook(excel_file_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == barcode:
                row[2].value = out_datetime
                row[3].value = in_datetime
                break

        # 엑셀 파일 저장
        wb.save(excel_file_path)

        # 데이터프레임 갱신
        global dF
        dF = load_students()

        return redirect(url_for('class_page', class_name=class_name))

    return render_template('class.html', students=filtered_df.to_dict(orient='records'), class_name=class_name)


@app.route('/manage_requests', methods=['GET'])
def manage_requests():
    if 'username' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    user_class = session['user_class']
    requests = OutingRequest.query.filter_by(student_class=user_class, approved=False).all()
    return render_template('manage_requests.html', requests=requests)


@app.route('/approve_request/<int:request_id>', methods=['POST'])
def approve_request(request_id):
    if 'username' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    outing_request = OutingRequest.query.get(request_id)
    if outing_request:
        outing_request.approved = True
        db.session.commit()

        # 엑셀 파일 수정
        wb = load_workbook(excel_file_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == outing_request.barcode:
                row[2].value = outing_request.out_time
                row[3].value = outing_request.in_time
                break

        # 엑셀 파일 저장
        wb.save(excel_file_path)

        # 데이터프레임 갱신
        global DF
        DF = load_students()

    return redirect(url_for('manage_requests'))


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)
