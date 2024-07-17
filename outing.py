from flask import Blueprint, request, render_template, redirect, url_for, flash, session
from models import db, User, OutingRequest
from openpyxl import load_workbook
import pandas as pd

outing_bp = Blueprint('outing', __name__)

excel_file_path = 'students.xlsx'

def load_students():
    """엑셀 파일을 읽고 데이터프레임을 반환합니다."""
    return pd.read_excel(excel_file_path, dtype={'barcode': str})

df = load_students()

@outing_bp.route('/student_home', methods=['GET'])
def student_home():
    if 'username' not in session or session['role'] != 'student':
        return redirect(url_for('login.login'))

    student = User.query.filter_by(username=session['username']).first()
    return render_template('student_home.html', student=student)

@outing_bp.route('/request_outing', methods=['POST'])
def request_outing():
    if 'username' not in session or session['role'] != 'student':
        return redirect(url_for('login.login'))

    student = User.query.filter_by(username=session['username']).first()
    out_time = request.form.get('out_time').replace('T', ' ')
    in_time = request.form.get('in_time').replace('T', ' ')

    outing_request = OutingRequest(
        student_id=student.id,
        student_name=student.username,
        student_class=student.user_class,
        barcode=student.barcode,
        out_time=out_time,
        in_time=in_time
    )

    db.session.add(outing_request)
    db.session.commit()

    flash('외출 신청이 완료되었습니다.', 'success')
    return redirect(url_for('outing.student_home'))

@outing_bp.route('/class/<class_name>', methods=['GET', 'POST'])
def class_page(class_name):
    global df  # global 선언이 df 변수 사용 전에 위치하도록 이동

    if 'username' not in session or session['role'] != 'teacher':
        return redirect(url_for('login.login'))

    filtered_df = df[df['class'] == class_name]
    if request.method == 'POST':
        barcode = request.form.get('barcode')
        out_datetime = request.form.get('out_datetime').replace('T', ' ')
        in_datetime = request.form.get('in_datetime').replace('T', ' ')

        # 엑셀 파일 수정
        wb = load_workbook(excel_file_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == barcode:
                row[2].value = out_datetime
                row[3].value = in_datetime
                break

        # 엑셀 파일 저장
        wb.save(excel_file_path)

        # 데이터프레임 갱신
        df = load_students()

        return redirect(url_for('outing.class_page', class_name=class_name))

    return render_template('class.html', students=filtered_df.to_dict(orient='records'), class_name=class_name)

@outing_bp.route('/manage_requests', methods=['GET'])
def manage_requests():
    if 'username' not in session or session['role'] != 'teacher':
        return redirect(url_for('login.login'))

    user_class = session['user_class']
    requests = OutingRequest.query.filter_by(student_class=user_class, approved=False).all()
    return render_template('manage_requests.html', requests=requests)

@outing_bp.route('/approve_request/<int:request_id>', methods=['POST'])
def approve_request(request_id):
    global df  # global 선언이 df 변수 사용 전에 위치하도록 이동

    if 'username' not in session or session['role'] != 'teacher':
        return redirect(url_for('login.login'))

    outing_request = OutingRequest.query.get(request_id)
    if outing_request:
        outing_request.approved = True
        db.session.commit()

        # 엑셀 파일 수정
        wb = load_workbook(excel_file_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == outing_request.barcode:
                row[2].value = outing_request.out_time
                row[3].value = outing_request.in_time
                break

        # 엑셀 파일 저장
        wb.save(excel_file_path)

        # 데이터프레임 갱신
        df = load_students()

    return redirect(url_for('outing.manage_requests'))
